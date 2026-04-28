import asyncio
import io
import json
import time
import logging
from typing import Any, List, Optional
from collections import defaultdict

from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Request, Depends, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    ALLOWED_ORIGINS,
    ANTHROPIC_API_KEY,
    RATE_LIMIT_PER_MINUTE,
    API_TOKEN,
    REQUIRE_API_TOKEN,
    CLAUDE_PROMPT_CACHE,
    MAX_CONCURRENT_DEFAULT,
    MAX_CONCURRENT_LIMIT,
    MAX_FILE_SIZE_MB,
)
from models import ExportItem, ExportExcelRequest
from extractors import extract_text, MAX_TEXT_LENGTH
from claude import call_claude, close_client, MODEL, CLAUDE_TWO_PASS
from user_errors import scoring_error_for_user

logger = logging.getLogger(__name__)

app = FastAPI(title="CV Scorer API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_event():
    await close_client()


_rate_store: dict[str, list[float]] = defaultdict(list)


async def check_rate_limit(request: Request):
    client_ip = request.client.host if request.client else "unknown"
    now = time.time()
    window = 60.0
    _rate_store[client_ip] = [t for t in _rate_store[client_ip] if now - t < window]
    if len(_rate_store[client_ip]) >= RATE_LIMIT_PER_MINUTE:
        raise HTTPException(
            status_code=429,
            detail=f"Rate limit dépassé ({RATE_LIMIT_PER_MINUTE} req/min). Réessayez plus tard.",
        )
    _rate_store[client_ip].append(now)


async def check_auth(request: Request):
    if REQUIRE_API_TOKEN:
        if not API_TOKEN:
            logger.error("REQUIRE_API_TOKEN activé mais API_TOKEN vide")
            raise HTTPException(
                status_code=503,
                detail="Serveur mal configuré (authentification requise).",
            )
        auth = request.headers.get("Authorization", "")
        if auth != f"Bearer {API_TOKEN}":
            raise HTTPException(status_code=401, detail="Token d'authentification invalide")
        return
    if not API_TOKEN:
        return
    auth = request.headers.get("Authorization", "")
    if auth != f"Bearer {API_TOKEN}":
        raise HTTPException(status_code=401, detail="Token d'authentification invalide")


@app.on_event("startup")
async def startup_event():
    if not ANTHROPIC_API_KEY:
        logger.critical("ANTHROPIC_API_KEY is not set!")
    if REQUIRE_API_TOKEN and not API_TOKEN:
        logger.critical("REQUIRE_API_TOKEN=1 mais API_TOKEN non défini — les routes /api/* renverront 503.")


@app.get("/health")
async def health():
    return {
        "status": "ok",
        "model": MODEL,
        "api_key_set": bool(ANTHROPIC_API_KEY),
        "claude_two_pass": CLAUDE_TWO_PASS,
        "claude_prompt_cache": CLAUDE_PROMPT_CACHE,
        "cv_text_max_chars": MAX_TEXT_LENGTH,
        "require_api_token": REQUIRE_API_TOKEN,
        "auth_configured": bool(API_TOKEN),
    }


@app.post("/api/score-stream", dependencies=[Depends(check_rate_limit), Depends(check_auth)])
async def score_stream(
    files: List[UploadFile] = File(...),
    poste: str = Form(...),
    max_concurrent: int = Form(default=MAX_CONCURRENT_DEFAULT),
    processing_mode: str = Form(default="parallel"),
):
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY non configurée")
    if not files:
        raise HTTPException(status_code=400, detail="Aucun fichier reçu")
    if not poste.strip():
        raise HTTPException(
            status_code=400,
            detail="La description du poste est obligatoire : sans ce texte, l’analyse ne peut pas comparer les CV au besoin.",
        )

    if processing_mode == "sequential":
        max_concurrent = 1
    else:
        max_concurrent = max(1, min(max_concurrent, MAX_CONCURRENT_LIMIT))

    file_data = []
    max_bytes = MAX_FILE_SIZE_MB * 1024 * 1024
    for f in files:
        name = (f.filename or "").strip()
        if not name.lower().endswith(".pdf"):
            raise HTTPException(
                status_code=400,
                detail=f"Format non accepté pour « {name or '(sans nom)'} » : seuls les fichiers PDF (.pdf) sont autorisés.",
            )
        content = await f.read()
        if len(content) > max_bytes:
            raise HTTPException(
                status_code=400,
                detail=f"Fichier {f.filename} trop volumineux ({len(content) // (1024*1024)}Mo > {MAX_FILE_SIZE_MB}Mo)",
            )
        file_data.append({"name": name, "content": content})

    async def event_generator():
        semaphore = asyncio.Semaphore(max_concurrent)
        results = []
        total = len(file_data)
        output_queue: asyncio.Queue = asyncio.Queue()
        runner: Optional[asyncio.Task] = None

        async def process_one(idx: int, fd: dict):
            async with semaphore:
                name = fd["name"]
                await output_queue.put(
                    f"data: {json.dumps({'type':'start','index':idx,'name':name,'total':total})}\n\n"
                )
                try:
                    t0 = time.perf_counter()
                    text = await extract_text(name, fd["content"])
                    extract_ms = (time.perf_counter() - t0) * 1000
                    logger.info(
                        "pipeline_metric event=extract_ok cv=%s extract_ms=%.1f text_chars=%d",
                        name,
                        extract_ms,
                        len(text or ""),
                    )
                    t1 = time.perf_counter()
                    result = await call_claude(text, name, poste)
                    claude_ms = (time.perf_counter() - t1) * 1000
                    logger.info(
                        "pipeline_metric event=claude_ok cv=%s claude_ms=%.1f two_pass=%s",
                        name,
                        claude_ms,
                        CLAUDE_TWO_PASS,
                    )
                    result["_file"] = name
                    result["_index"] = idx
                    results.append(result)
                    await output_queue.put(
                        f"data: {json.dumps({'type':'result','index':idx,'name':name,'data':result})}\n\n"
                    )
                except Exception as e:
                    logger.error("Error processing %s: %s", name, e, exc_info=True)
                    err_code, user_msg = scoring_error_for_user(e)
                    err = {
                        "_file": name,
                        "_index": idx,
                        "_error": user_msg,
                        "_error_code": err_code,
                        "score": 0,
                        "nom": name,
                        "recommandation": "Erreur d'analyse",
                        "decision": "non",
                        "profil_geographique": "inconnu",
                    }
                    results.append(err)
                    await output_queue.put(
                        f"data: {json.dumps({'type':'error','index':idx,'name':name,'error':user_msg,'error_code':err_code})}\n\n"
                    )

        async def run_all():
            tasks = [
                asyncio.create_task(process_one(i, fd))
                for i, fd in enumerate(file_data)
            ]
            await asyncio.gather(*tasks, return_exceptions=True)
            await output_queue.put(None)

        runner = asyncio.create_task(run_all())
        try:
            while True:
                item = await output_queue.get()
                if item is None:
                    break
                yield item

            await runner

            sorted_results = sorted(results, key=lambda x: x.get("score", 0), reverse=True)
            yield f"data: {json.dumps({'type':'complete','results':sorted_results})}\n\n"
        except asyncio.CancelledError:
            if runner and not runner.done():
                runner.cancel()
                try:
                    await runner
                except asyncio.CancelledError:
                    pass
            raise
        except Exception as e:
            logger.exception("score_stream interrompu: %s", e)
            if runner and not runner.done():
                runner.cancel()
                try:
                    await runner
                except (asyncio.CancelledError, Exception):
                    pass
            yield f"data: {json.dumps({'type': 'fatal', 'error': str(e)})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


def _rows_for_excel(req: ExportExcelRequest) -> List[ExportItem]:
    """Candidats avec score >= min_score, toutes décisions confondues, tri décroissant."""
    rows = [r for r in req.results if r.score >= req.min_score]
    rows.sort(key=lambda x: x.score, reverse=True)
    return rows


def _parse_export_body(body: Any) -> ExportExcelRequest:
    if isinstance(body, list):
        return ExportExcelRequest(results=[ExportItem.model_validate(x) for x in body])
    return ExportExcelRequest.model_validate(body)


def _profil_geographique_label(code: str) -> str:
    return {
        "national_tchad": "National (Tchad)",
        "international": "International",
        "inconnu": "—",
    }.get(code or "inconnu", "—")


def _build_export_sheet(ws, rows: List[ExportItem]) -> None:
    """Mise en forme sans Table Excel ni volets figés : le gel de la ligne 1 provoquait
    une double en-tête visible au défilement dans plusieurs versions d’Excel."""
    headers = [
        "Rang",
        "Nom",
        "Email",
        "Téléphone",
        "Score",
        "Profil géographique",
        "Niveau",
        "Années exp.",
        "Postes occupés",
        "Diplômes",
        "Points forts",
        "Points faibles",
        "Compétences clés",
        "Recommandation",
        "Fichier",
    ]
    ncols = len(headers)

    header_font = Font(bold=True, size=11, color="111827")
    body_font = Font(size=11, color="374151")
    header_fill = PatternFill(fill_type="solid", fgColor="F3F4F6")
    stripe_fill = PatternFill(fill_type="solid", fgColor="F9FAFB")
    thin = Side(style="thin", color="E5E7EB")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)
    wrap_center = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
        cell.border = grid_border

    center_cols = {1, 5, 6, 7, 8}  # rang, score, profil, niveau, années

    for i, r in enumerate(rows, 1):
        row_vals = [
            i,
            r.nom,
            r.email or "",
            r.telephone or "",
            r.score,
            _profil_geographique_label(r.profil_geographique),
            r.niveau,
            r.annees_experience if r.annees_experience is not None else "",
            "\n".join(r.postes_occupes),
            "\n".join(r.diplomes),
            " | ".join(r.points_forts),
            " | ".join(r.points_faibles),
            " | ".join(r.competences_cles),
            r.recommandation,
            r.file,
        ]
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=i + 1, column=c, value=val)
            cell.font = body_font
            cell.border = grid_border
            cell.alignment = wrap_center if c in center_cols else wrap

    last_row = 1 + len(rows)
    # Bandes alternées
    for row_idx in range(2, last_row + 1):
        if (row_idx - 2) % 2 == 1:
            for c in range(1, ncols + 1):
                ws.cell(row=row_idx, column=c).fill = stripe_fill

    ws.sheet_view.showGridLines = False

    widths = {
        "A": 8,   # Rang
        "B": 28,  # Nom
        "C": 32,  # Email
        "D": 16,  # Téléphone
        "E": 9,   # Score
        "F": 26,  # Profil géographique
        "G": 14,  # Niveau
        "H": 12,  # Années exp.
        "I": 36,  # Postes occupés
        "J": 32,  # Diplômes
        "K": 42,  # Points forts
        "L": 42,  # Points faibles
        "M": 36,  # Compétences clés
        "N": 48,  # Recommandation
        "O": 36,  # Fichier
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    ws.row_dimensions[1].height = 32
    for r in range(2, last_row + 1):
        ws.row_dimensions[r].height = 64

    # Filtres automatiques sur la ligne d’en-tête (Excel)
    last_col_letter = get_column_letter(ncols)
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"


@app.post("/api/export-excel", dependencies=[Depends(check_auth)])
async def export_excel(body: Any = Body(...)):
    """Export Excel : feuille complète (score ≥ min) + feuille Top N (sans filtre sur la décision)."""
    req = _parse_export_body(body)
    all_rows = _rows_for_excel(req)
    top_rows = all_rows[: req.top_n]

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Candidats"
    _build_export_sheet(ws_all, all_rows)

    ws_top = wb.create_sheet("Top 10")
    _build_export_sheet(ws_top, top_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="export_candidats.xlsx"',
        },
    )
